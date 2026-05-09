import openpyxl, sys, unicodedata


def print_menu():
    print("""
Επιλέξτε
1. Αναζήτηση μονάδας
2. Αναζήτηση γραφείου
3. Γενική αναζήτηση
4. Αναζήτηση μονάδας + γραφείου
5. Αναζήτηση τηλεφώνου
""")


def get_category():
    while True:
        category = input().strip()
        if category == "exit":
            sys.exit()
        if category in ("1", "2", "3", "4", "5"):
            return category
        print("Μη έγκυρη επιλογή.")


def normalize_greek(text):
    text = str(text).strip().casefold()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = "".join(text.split())
    return text


def matches_all_words(text, query):
    normalized_text = normalize_greek(text)
    words = query.strip().split()
    normalized_words = [normalize_greek(word) for word in words]
    return all(word in normalized_text for word in normalized_words)


def search_excel(excel_file, category, query=None, unit_query=None, office_query=None):
    found = False
    total_results = 0

    for sheet in excel_file.worksheets:
        for col in range(1, sheet.max_column + 1, 2):
            unit = sheet.cell(row=1, column=col).value

            if category == "1":
                if matches_all_words(unit, query):
                    print(f"\n=== {unit} ===")

                    for row in range(2, sheet.max_row + 1):
                        name = sheet.cell(row=row, column=col).value
                        phone = sheet.cell(row=row, column=col + 1).value

                        if name:
                            print(f"{name:<50} {phone}")
                            found = True
                    total_results += 1 

            elif category == "2":
                for row in range(2, sheet.max_row + 1):
                    name = sheet.cell(row=row, column=col).value

                    if name and matches_all_words(name, query):
                        phone = sheet.cell(row=row, column=col + 1).value
                        print("=" * 30)
                        print(f"{sheet.title} | {unit} -> {name} -> {phone}")
                        total_results += 1 
                        found = True

            elif category == "3":
                unit_found = False

                if unit and matches_all_words(unit, query):
                    print(f"\n=== {unit} ===")

                    for row in range(2, sheet.max_row + 1):
                        name = sheet.cell(row=row, column=col).value
                        phone = sheet.cell(row=row, column=col + 1).value

                        if name:
                            print(f"{name:<50} {phone}")
                            found = True
                            unit_found = True
                    total_results += 1 

                if not unit_found:
                    for row in range(2, sheet.max_row + 1):
                        name = sheet.cell(row=row, column=col).value

                        if matches_all_words(name, query):
                            phone = sheet.cell(row=row, column=col + 1).value
                            print("=" * 30)
                            print(f"{sheet.title} | {unit} -> {name} -> {phone}")
                            total_results += 1 
                            found = True

            elif category == "4" and unit:
                if matches_all_words(unit, unit_query):
                    for row in range(2, sheet.max_row + 1):
                        name = sheet.cell(row=row, column=col).value
                        phone = sheet.cell(row=row, column=col + 1).value

                        if name and matches_all_words(name, office_query):
                            print("=" * 30)
                            print(f"{unit} -> {name} -> {phone}")
                            total_results += 1 
                            found = True
            elif category == "5":
                for row in range(2, sheet.max_row + 1):
                    phone = sheet.cell(row=row, column=col + 1).value

                    if str(phone).strip() == query:
                        name = sheet.cell(row=row, column=col).value
                        print("=" * 30)
                        print(f"{sheet.title} | {unit} -> {name}")
                        total_results += 1
                        found = True

    return found, total_results


def main():
    excel_file = openpyxl.load_workbook("sample_data.xlsx")
    print("Πληκτρολογήστε 'exit' για τερματισμό.")

    while True:
        
        print_menu()
        category = get_category()

        if category == "4":
            unit_query = input("Μονάδα: ").strip()
            office_query = input("Γραφείο: ").strip()
            found, total_results = search_excel(excel_file, category, unit_query=unit_query, office_query=office_query)
        else:
            query = input("Αναζήτηση: ").strip()
            found, total_results = search_excel(excel_file, category, query=query)

        if not found:
            print("Δεν βρέθηκαν αποτελέσματα")
        else:
            print(f"\nΣύνολο αποτελεσμάτων: {total_results}")


if __name__ == "__main__":
    main()